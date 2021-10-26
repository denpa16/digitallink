from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Phone, File
from .serializers import PhoneSerializer, FileSerializer
from rest_framework.parsers import FileUploadParser, MultiPartParser, FormParser
from rest_framework import status
from openpyxl import load_workbook

# Create your views here.
class PhoneView(APIView):

    def post(self, request):
        data = request.data
        try:
            phone = Phone.objects.get(phone_number=data['phone_number'])
            return Response({"found": 1})
        except Phone.DoesNotExist:
            return Response({"found": 0})

class PhoneAddView(APIView):

    def post(self, request):
        phone = request.data
        found = request.data["found"]
        serializer = PhoneSerializer(data=phone)
        if serializer.is_valid(raise_exception=True):
            phone_saved = serializer.save()
        return Response({"success": "Phone '{}' added in stop list successfully".format(phone_saved.phone_number)})

class FileUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser, FileUploadParser)

    def post(self, request, *args, **kwargs):
        file_serializer = FileSerializer(data=request.data)
        if file_serializer.is_valid():
            file_serializer.save()
            obj_path = File.objects.get(id=file_serializer.instance.id).file.path
            wb = load_workbook(obj_path)
            sheet = wb.get_sheet_by_name('Sheet1')
            for i in range(1,sheet.max_row):
                phone = sheet.cell(row=i+1, column=1).value
                found = sheet.cell(row=i+1, column=2).value
                if found == 1:
                    data = {"phone_number": phone}
                    serializer = PhoneSerializer(data=data)
                    if serializer.is_valid(raise_exception=False):
                        serializer.save()
                        return Response(file_serializer.data, status=status.HTTP_201_CREATED)
                    else:
                        pass
        return Response(file_serializer.data, status=status.HTTP_201_CREATED)